# excel_filler.py
from openpyxl import load_workbook, Workbook

def fill_excel(template_path: str, data: dict, output_path: str):

    # Try loading template, create new if not found
    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Solar Analysis"

        # Create headers
        ws["A1"] = "Energybae — Solar Load Calculator"
        ws["B3"] = "Consumer Name"
        ws["B4"] = "Consumer No"
        ws["B5"] = "Connection Type"
        ws["B6"] = "Sanctioned Load (kW)"
        ws["B7"] = "Fixed Charges"
        ws["B8"] = "Solar Panel Wattage"
        ws["C8"] = 600
        ws["B10"] = "Sr.No"
        ws["C10"] = "Month"
        ws["D10"] = "Units"
        ws["E10"] = "Bill Amount"
        ws["F10"] = "Unit Cost"

    # Fill customer details
    ws["D1"] = data.get("consumer_name", "")
    ws["D2"] = data.get("consumer_no", "")
    ws["D4"] = data.get("sanctioned_load_kw", "")
    ws["D5"] = data.get("connection_type", "")

    # Fill monthly units (D9 to D20)
    monthly_units = data.get("monthly_units", [])
    start_row = 9
    for i in range(12):
        row = start_row + i
        if i < len(monthly_units):
            ws[f"D{row}"] = monthly_units[i].get("units", 0)
        else:
            ws[f"D{row}"] = ""

    # Fill bill amount
    ws["E20"] = data.get("bill_amount", 0)

    # Save with consumer name + bill month as filename
    consumer = data.get("consumer_name", "Customer").replace(" ", "_")
    month = data.get("bill_month", "").replace(" ", "_").replace("-", "_")
    output_path = f"{consumer}_{month}_Solar.xlsx"

    wb.save(output_path)
    return output_path